"""FastAPI integration for Odoo-PostgreSQL sync service.

This module provides a REST API for managing synchronization tasks.
Designed for production deployment with uvicorn.
"""

from datetime import date, datetime
from io import BytesIO
import base64
import hashlib
import hmac
import json
from decimal import Decimal
from html import escape
from pathlib import Path
from typing import Optional
from urllib.parse import urlencode

from fastapi import BackgroundTasks, Depends, FastAPI, Form, HTTPException, Query, Request, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy import bindparam, text

from src.engine.sync_engine import SyncEngine
from src.engine.scheduler import SyncScheduler
from src.state.state_manager import StateManager
from src.clients.postgres_client import PostgresClient
from src.utils.logging import get_logger, setup_logging
from src.utils.settings import get_settings

# Initialize logging
setup_logging()
logger = get_logger("api")

APP_SETTINGS = get_settings()
DASHBOARD_USERNAME = APP_SETTINGS.dashboard_username
DASHBOARD_PASSWORD = APP_SETTINGS.dashboard_password
SESSION_SECRET = APP_SETTINGS.session_secret or "dashboard-dev-session-secret"
if not APP_SETTINGS.session_secret:
    logger.warning("SESSION_SECRET missing; using local development fallback secret for demo auth.")

DEFAULT_DASHBOARD_PATH = "/dashboard/internal-order-rekap"
PROTECTED_PAGE_PATHS = {
    "/",
    "/dashboard/internal-orders",
    "/dashboard/sales-orders",
    "/dashboard/internal-order-rekap",
    "/dashboard/control-tower",
}
PROTECTED_API_PREFIX = "/api/dashboard/"
DASHBOARD_SESSION_COOKIE = "dashboard_session"
DASHBOARD_SESSION_TTL_SECONDS = 60 * 60 * 24 * 7

# Create FastAPI app
app = FastAPI(
    title="Odoo-PostgreSQL Sync API",
    description="REST API for Odoo to PostgreSQL synchronization",
    version="2.0.0",
)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
STATIC_DIR = Path(__file__).resolve().parent / "static"
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

def sign_dashboard_session(payload: dict) -> str:
    raw = json.dumps(payload, separators=(",", ":"), sort_keys=True).encode("utf-8")
    encoded = base64.urlsafe_b64encode(raw).decode("ascii").rstrip("=")
    signature = hmac.new(SESSION_SECRET.encode("utf-8"), encoded.encode("ascii"), hashlib.sha256).hexdigest()
    return f"{encoded}.{signature}"


def read_dashboard_session(raw_value: Optional[str]) -> Optional[dict]:
    if not raw_value or "." not in raw_value:
        return None
    encoded, signature = raw_value.rsplit(".", 1)
    expected = hmac.new(SESSION_SECRET.encode("utf-8"), encoded.encode("ascii"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(signature, expected):
        return None
    padded = encoded + "=" * (-len(encoded) % 4)
    try:
        decoded = base64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
        payload = json.loads(decoded)
    except (ValueError, json.JSONDecodeError, UnicodeDecodeError):
        return None
    if isinstance(payload, dict) and payload.get("dashboard_authenticated"):
        return payload
    return None


def is_authenticated(request: Request) -> bool:
    return read_dashboard_session(request.cookies.get(DASHBOARD_SESSION_COOKIE)) is not None


def safe_next_path(value: Optional[str]) -> str:
    candidate = (value or "").strip()
    if candidate in PROTECTED_PAGE_PATHS and candidate != "/":
        return candidate
    return DEFAULT_DASHBOARD_PATH


def render_login_page(
    error: Optional[str] = None,
    next_path: str = DEFAULT_DASHBOARD_PATH,
    username: str = "",
) -> str:
    error_html = f'<div class="login-error">{escape(error)}</div>' if error else ""
    next_value = escape(next_path)
    username_value = escape(username)
    template = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Dashboard Login</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f4f7f8;
      --surface: #ffffff;
      --border: #d9e1e5;
      --text: #17212b;
      --muted: #66737f;
      --accent: #147a78;
      --accent-strong: #0f6160;
      --danger: #a93c32;
      --shadow: 0 12px 32px rgba(23, 33, 43, 0.10);
      --radius: 12px;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      min-height: 100vh;
      font-family: Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      background: var(--bg);
      color: var(--text);
      display: grid;
      place-items: center;
      padding: 24px;
    }
    .login-shell {
      width: min(100%, 440px);
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: var(--radius);
      box-shadow: var(--shadow);
      padding: 28px;
    }
    .login-eyebrow {
      font-size: 11px;
      font-weight: 760;
      letter-spacing: 0;
      text-transform: uppercase;
      color: var(--muted);
      margin: 0 0 8px;
    }
    h1 {
      margin: 0 0 8px;
      font-size: 26px;
      line-height: 1.15;
    }
    p {
      margin: 0 0 18px;
      color: var(--muted);
      line-height: 1.5;
    }
    form {
      display: grid;
      gap: 14px;
    }
    label {
      display: grid;
      gap: 6px;
      font-size: 12px;
      font-weight: 650;
      color: var(--muted);
    }
    input {
      height: 40px;
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 0 12px;
      font: inherit;
      color: var(--text);
      background: #fff;
    }
    input:focus {
      outline: 2px solid rgba(20, 122, 120, 0.18);
      outline-offset: 1px;
      border-color: rgba(20, 122, 120, 0.35);
    }
    .login-error {
      border: 1px solid rgba(169, 60, 50, 0.25);
      background: #fff4f2;
      color: var(--danger);
      border-radius: 10px;
      padding: 12px 14px;
      font-size: 13px;
      font-weight: 600;
      margin-bottom: 14px;
    }
    .login-actions {
      display: flex;
      justify-content: flex-end;
      gap: 10px;
      margin-top: 2px;
    }
    .primary-button {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-height: 40px;
      border: 1px solid var(--accent);
      border-radius: 10px;
      padding: 0 16px;
      background: var(--accent);
      color: #fff;
      cursor: pointer;
      font: inherit;
      font-weight: 700;
    }
    .primary-button:hover { background: var(--accent-strong); }
    .hint {
      margin-top: 16px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.45;
    }
  </style>
</head>
<body>
  <main class="login-shell">
    <div class="login-eyebrow">Local Demo Access</div>
    <h1>Dashboard Login</h1>
    <p>Use the local demo credentials to open the dashboards.</p>
    __ERROR_HTML__
    <form method="post" action="/login">
      <input type="hidden" name="next" value="__NEXT_VALUE__">
      <label>
        <span>Username</span>
        <input name="username" type="text" value="__USERNAME_VALUE__" autocomplete="username" required autofocus>
      </label>
      <label>
        <span>Password</span>
        <input name="password" type="password" autocomplete="current-password" required>
      </label>
      <div class="login-actions">
        <button class="primary-button" type="submit">Sign in</button>
      </div>
    </form>
    <div class="hint">Default destination after login is the Internal Order Rekap dashboard.</div>
  </main>
</body>
</html>"""
    return (
        template.replace("__ERROR_HTML__", error_html)
        .replace("__NEXT_VALUE__", next_value)
        .replace("__USERNAME_VALUE__", username_value)
    )


def redirect_to_login(next_path: str = DEFAULT_DASHBOARD_PATH) -> RedirectResponse:
    return RedirectResponse(
        url=f"/login?{urlencode({'next': next_path})}",
        status_code=status.HTTP_307_TEMPORARY_REDIRECT,
    )


@app.middleware("http")
async def dashboard_auth_middleware(request: Request, call_next):
    path = request.url.path

    if path.startswith("/static") or path in {"/login", "/logout"} or path.startswith("/docs") or path.startswith("/redoc") or path.startswith("/openapi.json"):
        return await call_next(request)

    if path == "/":
        if is_authenticated(request):
            return RedirectResponse(url=DEFAULT_DASHBOARD_PATH, status_code=status.HTTP_307_TEMPORARY_REDIRECT)
        return redirect_to_login(DEFAULT_DASHBOARD_PATH)

    if path in PROTECTED_PAGE_PATHS:
        if not is_authenticated(request):
            return redirect_to_login(path)

    if path.startswith(PROTECTED_API_PREFIX):
        if not is_authenticated(request):
            return JSONResponse(status_code=status.HTTP_401_UNAUTHORIZED, content={"detail": "Authentication required."})

    return await call_next(request)


@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@app.get("/login", include_in_schema=False)
async def login_page(request: Request, next: Optional[str] = None):
    """Render the local dashboard login page."""
    if is_authenticated(request):
        return RedirectResponse(url=DEFAULT_DASHBOARD_PATH, status_code=status.HTTP_303_SEE_OTHER)
    return HTMLResponse(render_login_page(next_path=safe_next_path(next)))


@app.post("/login", include_in_schema=False)
async def login_submit(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    next: Optional[str] = Form(default=None),
):
    """Authenticate the local demo user and create a signed session."""
    next_path = safe_next_path(next)
    if username == DASHBOARD_USERNAME and password == DASHBOARD_PASSWORD:
        response = RedirectResponse(url=next_path, status_code=status.HTTP_303_SEE_OTHER)
        response.set_cookie(
            key=DASHBOARD_SESSION_COOKIE,
            value=sign_dashboard_session({"dashboard_authenticated": True, "dashboard_username": username}),
            max_age=DASHBOARD_SESSION_TTL_SECONDS,
            httponly=True,
            secure=False,
            samesite="lax",
            path="/",
        )
        return response

    return HTMLResponse(
        render_login_page(
            error="Invalid username or password.",
            next_path=next_path,
            username=username,
        ),
        status_code=status.HTTP_401_UNAUTHORIZED,
    )


@app.get("/logout", include_in_schema=False)
async def logout(request: Request):
    """Clear the dashboard session and return to login."""
    response = RedirectResponse(url="/login", status_code=status.HTTP_303_SEE_OTHER)
    response.delete_cookie(DASHBOARD_SESSION_COOKIE, path="/")
    return response


# Global instances
_scheduler: Optional[SyncScheduler] = None


# Pydantic models for API
class SyncRequest(BaseModel):
    """Request model for sync operation."""

    full_sync: bool = Field(default=False, description="Perform full sync")
    models: Optional[list[str]] = Field(default=None, description="Specific models to sync")


class SyncResponse(BaseModel):
    """Response model for sync operation."""

    status: str
    message: str
    results: Optional[list[dict]] = None
    timestamp: datetime = Field(default_factory=datetime.utcnow)


class StatusResponse(BaseModel):
    """Response model for status check."""

    status: str
    models: list[dict]
    total_models: int
    synced_models: int
    scheduler_running: bool


class HealthResponse(BaseModel):
    """Response model for health check."""

    status: str
    odoo_connected: bool
    postgres_connected: bool
    scheduler_running: bool
    timestamp: datetime = Field(default_factory=datetime.utcnow)


class SyncHistoryResponse(BaseModel):
    """Response model for sync history."""

    records: list[dict]
    total: int


class SyncAuditResponse(BaseModel):
    """Response model for sync audit."""

    records: list[dict]
    total: int


class ResetRequest(BaseModel):
    """Request model for reset operation."""

    models: list[str] = Field(..., description="Models to reset")


def _json_safe(value):
    """Convert database values into JSON-friendly primitives."""
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, list):
        return [_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    return value


def _calculate_ratio(numerator: float, denominator: float) -> Optional[float]:
    if not denominator:
        return None
    return numerator / denominator


def _build_internal_order_dashboard_summary(rows: list[dict]) -> dict:
    active_rows = [row for row in rows if row.get("traceability_status") != "CANCELLED_RECORD"]

    so_ordered = sum(float(row.get("total_so_ordered_qty") or 0) for row in active_rows)
    so_delivered = sum(float(row.get("total_so_delivered_qty") or 0) for row in active_rows)
    so_invoiced = sum(float(row.get("total_so_invoiced_qty") or 0) for row in active_rows)
    po_ordered = sum(float(row.get("total_po_ordered_qty") or 0) for row in active_rows)
    po_received = sum(float(row.get("total_po_received_qty") or 0) for row in active_rows)
    po_invoiced = sum(float(row.get("total_po_invoiced_qty") or 0) for row in active_rows)

    return {
        "active_internal_orders": len(active_rows),
        "internal_orders_with_mo": sum(1 for row in active_rows if row.get("linked_mo_count", 0) > 0),
        "internal_orders_with_so": sum(1 for row in active_rows if row.get("linked_so_count", 0) > 0),
        "internal_orders_delivered": sum(1 for row in active_rows if row.get("has_delivered_so")),
        "internal_orders_invoiced": sum(1 for row in active_rows if row.get("has_invoiced_so")),
        "delivery_progress_ratio": _calculate_ratio(so_delivered, so_ordered),
        "invoice_progress_ratio": _calculate_ratio(so_invoiced, so_ordered),
        "procurement_receipt_progress_ratio": _calculate_ratio(po_received, po_ordered),
        "procurement_billing_progress_ratio": _calculate_ratio(po_invoiced, po_ordered),
        "total_so_ordered_qty": so_ordered,
        "total_so_delivered_qty": so_delivered,
        "total_so_invoiced_qty": so_invoiced,
        "total_po_ordered_qty": po_ordered,
        "total_po_received_qty": po_received,
        "total_po_invoiced_qty": po_invoiced,
    }


def _build_internal_order_filter_options(rows: list[dict]) -> dict:
    def unique_values(key: str) -> list[str]:
        values = {
            str(row[key])
            for row in rows
            if row.get(key) is not None and str(row.get(key)).strip()
        }
        return sorted(values)

    return {
        "requesters": unique_values("requester"),
        "statuses": unique_values("status_summary"),
        "traceability_statuses": unique_values("traceability_status"),
    }


def _build_sales_order_dashboard_summary(rows: list[dict]) -> dict:
    active_rows = [row for row in rows if row.get("follow_up_status") != "CANCELLED_RECORD"]

    ordered_qty = sum(float(row.get("ordered_qty") or 0) for row in active_rows)
    delivered_qty = sum(float(row.get("delivered_qty") or 0) for row in active_rows)
    invoiced_qty = sum(float(row.get("invoiced_qty") or 0) for row in active_rows)
    ordered_amount = sum(float(row.get("ordered_amount") or 0) for row in active_rows)
    delivered_amount = sum(float(row.get("delivered_amount") or 0) for row in active_rows)
    invoiced_amount = sum(float(row.get("invoiced_amount") or 0) for row in active_rows)

    return {
        "active_sales_orders": len(active_rows),
        "delivered_sales_orders": sum(1 for row in active_rows if row.get("has_delivered_qty")),
        "invoiced_sales_orders": sum(1 for row in active_rows if row.get("has_invoiced_qty")),
        "delayed_delivery_sales_orders": sum(
            1 for row in active_rows if row.get("follow_up_status") == "DELAYED_DELIVERY"
        ),
        "waiting_invoice_sales_orders": sum(
            1 for row in active_rows if row.get("follow_up_status") == "WAITING_INVOICE"
        ),
        "quantity_delivery_progress_ratio": _calculate_ratio(delivered_qty, ordered_qty),
        "quantity_invoice_progress_ratio": _calculate_ratio(invoiced_qty, ordered_qty),
        "amount_delivery_progress_ratio": _calculate_ratio(delivered_amount, ordered_amount),
        "amount_invoice_progress_ratio": _calculate_ratio(invoiced_amount, ordered_amount),
        "sales_orders_from_internal_order": sum(
            1 for row in active_rows if row.get("source_type") == "FROM_INTERNAL_ORDER"
        ),
        "sales_orders_make_to_order": sum(
            1 for row in active_rows if row.get("source_type") == "MAKE_TO_ORDER"
        ),
        "sales_orders_from_stock": sum(
            1 for row in active_rows if row.get("source_type") == "FROM_STOCK"
        ),
        "unknown_source_sales_orders": sum(
            1 for row in active_rows if row.get("source_type") == "UNKNOWN_SOURCE"
        ),
        "total_ordered_qty": ordered_qty,
        "total_delivered_qty": delivered_qty,
        "total_invoiced_qty": invoiced_qty,
        "total_ordered_amount": ordered_amount,
        "total_delivered_amount": delivered_amount,
        "total_invoiced_amount": invoiced_amount,
    }


def _build_sales_order_filter_options(rows: list[dict]) -> dict:
    def unique_values(key: str) -> list[str]:
        values = {
            str(row[key])
            for row in rows
            if row.get(key) is not None and str(row.get(key)).strip()
        }
        return sorted(values)

    return {
        "customers": unique_values("customer_name"),
        "years": unique_values("order_year"),
        "product_types": unique_values("product_type_label"),
        "source_types": unique_values("source_type"),
        "sales_order_statuses": unique_values("sales_order_state"),
        "follow_up_statuses": unique_values("follow_up_status"),
    }


def _build_internal_order_rekap_warnings(summary_row: dict) -> list[str]:
    warnings: list[str] = []

    if not summary_row.get("has_sales_order_link"):
        warnings.append("Pre-SO Internal Order: no linked Sales Order found.")
    if int(summary_row.get("mixed_uom_count") or 0) > 0:
        warnings.append("Mixed UoM detected; quantity comparison may be unreliable.")
    if Decimal(str(summary_row.get("rkb_actual_unknown_class_amount") or 0)) != 0:
        warnings.append("Unknown product classification amount exists; review product classification.")
    if int(summary_row.get("po_without_rop_count") or 0) > 0:
        warnings.append("Some PO rows exist without linked ROP.")
    if int(summary_row.get("rop_without_po_count") or 0) > 0:
        warnings.append("Some ROP rows exist without linked PO.")

    return warnings


def get_sync_engine() -> SyncEngine:
    """Dependency to get sync engine instance."""
    engine = SyncEngine()
    engine.initialize()
    return engine


def get_scheduler() -> Optional[SyncScheduler]:
    """Get scheduler instance."""
    return _scheduler


# ============================================
# Health Endpoints
# ============================================

@app.get("/", include_in_schema=False)
async def dashboard_home():
    """Open the first dashboard page by default."""
    return FileResponse(STATIC_DIR / "dashboard" / "internal-orders.html")


@app.get("/dashboard/internal-orders", include_in_schema=False)
async def internal_order_dashboard_page():
    """Serve the Internal Order Traceability Dashboard page."""
    return FileResponse(STATIC_DIR / "dashboard" / "internal-orders.html")


@app.get("/dashboard/sales-orders", include_in_schema=False)
async def sales_order_dashboard_page():
    """Serve the Sales Order Traceability Dashboard page."""
    return FileResponse(STATIC_DIR / "dashboard" / "sales-orders.html")


@app.get("/dashboard/internal-order-rekap", include_in_schema=False)
async def internal_order_rekap_dashboard_page():
    """Serve the Internal Order Rekap Dashboard page."""
    return FileResponse(STATIC_DIR / "dashboard" / "internal-order-rekap.html")


@app.get("/dashboard/control-tower", include_in_schema=False)
async def control_tower_dashboard_page():
    """Serve the read-only Control Tower dashboard."""
    return FileResponse(STATIC_DIR / "dashboard" / "control-tower.html")


class DashboardExportColumn(BaseModel):
    key: str
    label: str
    type: str = Field(default="string")


class DashboardExportRequest(BaseModel):
    file_name: str
    sheet_name: str = Field(default="Sheet1")
    columns: list[DashboardExportColumn]
    rows: list[dict]


def _safe_excel_sheet_name(value: str) -> str:
    cleaned = "".join("_" if character in '[]:*?/\\' else character for character in (value or "Sheet1"))
    cleaned = cleaned.strip() or "Sheet1"
    return cleaned[:31]


def _safe_excel_file_name(value: str) -> str:
    cleaned = "".join(character if character.isalnum() or character in "._-" else "_" for character in (value or "dashboard.xlsx"))
    if not cleaned.lower().endswith(".xlsx"):
        cleaned = f"{cleaned}.xlsx"
    return cleaned or "dashboard.xlsx"


def _excel_cell_value(value, cell_type: str):
    if value is None or value == "":
        return None
    if cell_type in {"number", "currency", "percent"}:
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    if cell_type == "date":
        if isinstance(value, (datetime, date)):
            return value
        text_value = str(value).strip()
        if not text_value:
            return None
        try:
            if len(text_value) == 10:
                return date.fromisoformat(text_value)
            return datetime.fromisoformat(text_value.replace("Z", "+00:00"))
        except ValueError:
            return text_value
    return str(value)


def _build_dashboard_export_workbook(payload: DashboardExportRequest) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _safe_excel_sheet_name(payload.sheet_name)
    worksheet.freeze_panes = "A2"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="147A78")
    header_border = Border(bottom=Side(style="thin", color="D9E0E6"))
    header_alignment = Alignment(horizontal="center", vertical="center")
    body_alignment = Alignment(vertical="top")
    numeric_alignment = Alignment(horizontal="right", vertical="top")

    for column_index, column in enumerate(payload.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column.label)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_alignment

    for row_index, row in enumerate(payload.rows, start=2):
        for column_index, column in enumerate(payload.columns, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=_excel_cell_value(row.get(column.key), column.type))
            if column.type == "currency":
                cell.number_format = '#,##0.00'
                cell.alignment = numeric_alignment
            elif column.type == "percent":
                cell.number_format = '0.0%'
                cell.alignment = numeric_alignment
            elif column.type == "number":
                cell.number_format = '#,##0.00'
                cell.alignment = numeric_alignment
            elif column.type == "date" and isinstance(cell.value, (datetime, date)):
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = body_alignment
            else:
                cell.alignment = body_alignment

    worksheet.auto_filter.ref = worksheet.dimensions
    for column_index, column in enumerate(payload.columns, start=1):
        values = [column.label, *["" if row.get(column.key) is None else str(row.get(column.key)) for row in payload.rows[:200]]]
        width = max(len(value) for value in values) + 2
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(width, 12), 42)

    return workbook

@app.post("/api/dashboard/export/xlsx", tags=["Dashboard"])
async def dashboard_export_xlsx(payload: DashboardExportRequest):
    if not payload.columns:
        raise HTTPException(status_code=400, detail="At least one export column is required.")

    workbook = _build_dashboard_export_workbook(payload)
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{_safe_excel_file_name(payload.file_name)}\""},
    )

@app.get("/api/dashboard/internal-orders", tags=["Dashboard"])
async def internal_order_dashboard_data():
    """
    Return V1 Internal Order traceability dashboard data.

    This endpoint is read-only and uses the validated dashboard view:
    vw_dashboard_internal_order_traceability.
    """
    sql = text("""
        SELECT
            internal_order_number,
            traceability_status,
            status_summary,
            requester,
            needed_date_from,
            needed_date_to,
            line_count,
            product_count,
            linked_mo_count,
            linked_so_count,
            linked_so_line_count,
            total_so_amount,
            total_so_ordered_qty,
            total_so_delivered_qty,
            total_so_invoiced_qty,
            so_delivery_progress_ratio,
            so_invoice_progress_ratio,
            has_delivered_so,
            has_invoiced_so,
            delivery_status_summary,
            invoice_status_summary,
            linked_po_line_count,
            total_po_ordered_qty,
            total_po_received_qty,
            total_po_invoiced_qty,
            po_receipt_progress_ratio,
            po_invoice_progress_ratio,
            purchase_status_summary,
            accounting_line_count,
            manufacturing_movement_count,
            finished_goods_store_count,
            delivery_movement_count
        FROM vw_dashboard_internal_order_traceability
        ORDER BY
            CASE traceability_status
                WHEN 'HAS_ACCOUNTING_LINK' THEN 1
                WHEN 'HAS_INVOICED_SO' THEN 2
                WHEN 'HAS_DELIVERED_SO' THEN 3
                WHEN 'HAS_LINKED_SO' THEN 4
                WHEN 'HAS_MO_NO_SO_YET' THEN 5
                WHEN 'NEW_OR_TO_SUBMIT_NO_MO' THEN 6
                WHEN 'OLD_OR_UNLINKED_NO_MO' THEN 7
                WHEN 'CANCELLED_RECORD' THEN 8
                ELSE 9
            END,
            internal_order_number
    """)

    pg = PostgresClient()
    try:
        with pg.engine.connect() as conn:
            result = conn.execute(sql)
            rows = [
                {key: _json_safe(value) for key, value in row._mapping.items()}
                for row in result.fetchall()
            ]

        return {
            "rows": rows,
            "summary": _build_internal_order_dashboard_summary(rows),
            "filters": _build_internal_order_filter_options(rows),
            "meta": {
                "source_view": "vw_dashboard_internal_order_traceability",
                "row_count": len(rows),
                "profitability_included": False,
                "stock_movement_counts_are_diagnostic": True,
            },
        }
    except Exception as e:
        logger.error("Internal Order dashboard query failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        pg.close()


@app.get("/api/dashboard/sales-orders", tags=["Dashboard"])
async def sales_order_dashboard_data():
    """
    Return Phase 2A Sales Order traceability dashboard data.

    This endpoint is read-only and uses the dashboard view:
    vw_dashboard_sales_order_traceability.
    """
    sql = text("""
        SELECT
            sales_order_id,
            sales_order_number,
            company_id,
            customer_name,
            order_date,
            order_create_date,
            order_year,
            commitment_date,
            sales_order_state,
            normalized_status,
            is_cancelled,
            is_valid_for_metrics,
            delivery_status,
            invoice_status,
            product_type_raw,
            product_type_label,
            raw_internal_order_reference,
            source_type,
            source_link_status,
            from_internal_order_line_count,
            from_stock_line_count,
            make_to_order_line_count,
            mixed_source_line_count,
            unknown_source_line_count,
            sales_order_line_count,
            ordered_qty,
            delivered_qty,
            invoiced_qty,
            ordered_amount,
            delivered_amount,
            invoiced_amount,
            ordered_amount_idr,
            delivered_amount_idr,
            invoiced_amount_idr,
            sales_amount_idr,
            rkb_planned_cost,
            direct_rkb_planned_cost,
            io_correlated_rkb_planned_cost,
            io_correlated_rkb_planned_cost_full,
            io_rkb_planned_cost_per_unit,
            io_correlated_rkb_planned_cost_allocated,
            io_rkb_unallocated_planned_cost,
            rkb_allocation_basis,
            rkb_cost_basis,
            direct_actual_cost,
            direct_actual_cost_per_unit,
            direct_actual_cost_basis,
            io_backed_actual_cost,
            io_backed_actual_cost_full,
            io_backed_actual_cost_allocated,
            io_backed_actual_cost_per_unit,
            io_backed_actual_cost_basis,
            io_backed_actual_cost_is_correlation_only,
            total_related_actual_cost,
            total_related_actual_cost_full,
            actual_cost,
            actual_cost_quantity_based,
            actual_cost_per_unit,
            actual_cost_basis,
            rkb_kontribusi_amount,
            rkb_kontribusi_percent,
            kontribusi_aktual_amount,
            kontribusi_aktual_percent,
            contribution_basis_warning,
            currency_rate_used,
            currency_conversion_basis,
            qty_delivery_progress_ratio,
            qty_invoice_progress_ratio,
            amount_delivery_progress_ratio,
            amount_invoice_progress_ratio,
            is_fully_delivered_qty,
            is_fully_invoiced_qty,
            is_fully_delivered_amount,
            is_fully_invoiced_amount,
            has_delivered_qty,
            has_invoiced_qty,
            internal_order_count,
            manufacturing_order_count,
            job_order_mo_count,
            direct_mo_count,
            direct_mo_qty,
            direct_done_mo_qty,
            direct_in_progress_mo_qty,
            io_backed_mo_count,
            io_backed_mo_qty,
            io_backed_done_mo_qty,
            io_backed_in_progress_mo_qty,
            total_related_mo_count,
            total_related_mo_qty,
            total_done_mo_qty,
            total_in_progress_mo_qty,
            shared_io_count,
            shared_io_numbers,
            multi_io_so_count,
            has_multi_io_so,
            linked_so_qty_basis,
            io_qty_correlation_status,
            accounting_line_count,
            stock_movement_diagnostic_count,
            unknown_movement_diagnostic_count,
            follow_up_status,
            follow_up_status_priority,
            '[]'::jsonb AS sales_order_lines,
            '[]'::jsonb AS manufacturing_orders,
            '[]'::jsonb AS io_manufacturing_correlations,
            diagnostics
        FROM vw_dashboard_sales_order_traceability
        ORDER BY
            follow_up_status_priority,
            commitment_date NULLS LAST,
            sales_order_number
    """)

    pg = PostgresClient()
    try:
        with pg.engine.connect() as conn:
            result = conn.execute(sql)
            rows = [
                {key: _json_safe(value) for key, value in row._mapping.items()}
                for row in result.fetchall()
            ]

        return {
            "rows": rows,
            "summary": _build_sales_order_dashboard_summary(rows),
            "filters": _build_sales_order_filter_options(rows),
            "meta": {
                "source_view": "vw_dashboard_sales_order_traceability",
                "row_count": len(rows),
                "phase": "2A.1",
                "profitability_included": False,
                "accounting_ar_included": False,
                "stock_movement_counts_are_diagnostic": True,
                "io_mo_quantity_is_correlation_only": True,
                "contribution_metrics_included": True,
                "contribution_metrics_are_accounting_profit": False,
            },
        }
    except Exception as e:
        logger.error("Sales Order dashboard query failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        pg.close()





@app.get("/api/dashboard/sales-orders/{sales_order_id}/details", tags=["Dashboard"])
async def sales_order_dashboard_detail(sales_order_id: int):
    """Return allowed expanded detail sections for one Sales Order row."""
    sql = text("""
        SELECT
            sales_order_id,
            sales_order_number,
            sales_order_lines,
            manufacturing_orders,
            io_manufacturing_correlations
        FROM vw_dashboard_sales_order_traceability
        WHERE sales_order_id = :sales_order_id
    """)

    pg = PostgresClient()
    try:
        with pg.engine.connect() as conn:
            row = conn.execute(sql, {"sales_order_id": sales_order_id}).mappings().first()

        if row is None:
            raise HTTPException(status_code=404, detail="Sales Order not found")

        return {key: _json_safe(value) for key, value in row.items()}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Sales Order dashboard detail query failed", sales_order_id=sales_order_id, error=str(e))
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        pg.close()

@app.get("/api/dashboard/internal-order-rekap", tags=["Dashboard"])
async def internal_order_rekap_dashboard_data(
    internal_order_number: Optional[str] = Query(default=None),
    perspective: str = Query(default="internal_order"),
    sales_order_number: Optional[str] = Query(default=None),
):
    """
    Return Internal Order Rekap operational reconciliation data.

    This endpoint is read-only and uses:
    vw_internal_order_rekap_summary
    vw_internal_order_rekap_lines
    """
    normalized_perspective = (perspective or "internal_order").strip().lower().replace("-", "_")
    if normalized_perspective == "internal_order" and sales_order_number and not internal_order_number:
        normalized_perspective = "sales_order"
    if normalized_perspective not in {"internal_order", "sales_order"}:
        raise HTTPException(
            status_code=400,
            detail={"error": "perspective must be internal_order or sales_order."},
        )

    def _decimal_value(value) -> Decimal:
        if value is None or value == "":
            return Decimal("0")
        return Decimal(str(value))

    def _ratio_value(numerator: Decimal, denominator: Decimal):
        if denominator == 0:
            return None
        return numerator / denominator

    if normalized_perspective == "sales_order":
        if sales_order_number is None or not sales_order_number.strip():
            raise HTTPException(
                status_code=400,
                detail={"error": "sales_order_number query parameter is required for sales_order perspective."},
            )

        normalized_sales_order_number = sales_order_number.strip()
        context_note = "Sales Order Perspective shows linked IO and/or direct SO/JO material chain context. This is traceability context, not product-level allocation or accounting margin."

        selected_sales_order_sql = text("""
            SELECT
                so.id AS sales_order_id,
                so.name::text AS sales_order_number,
                so.state::text AS sales_order_state,
                (COALESCE(so.amount_untaxed, 0) * COALESCE(NULLIF(so.currency_rate, 0), 1))::numeric AS sales_order_amount
            FROM sale_order so
            WHERE BTRIM(so.name::text) = :sales_order_number
              AND COALESCE(so.state::text, '') NOT ILIKE '%cancel%'
            ORDER BY so.id DESC
            LIMIT 1
        """)

        linked_internal_orders_sql = text("""
            WITH internal_order_rows AS (
                SELECT DISTINCT
                    NULLIF(BTRIM(COALESCE(io.internal_order_number::text, '')), '') AS internal_order_number,
                    io.approval_request_numeric_id::bigint AS internal_order_id
                FROM vw_approval_product_line_context io
                WHERE io.approval_business_type = 'INTERNAL_ORDER'
                  AND UPPER(BTRIM(COALESCE(io.approval_category_raw::text, ''))) = 'MANUFACTURE'
                  AND io.is_valid_for_metrics
                  AND NULLIF(BTRIM(COALESCE(io.internal_order_number::text, '')), '') IS NOT NULL
                  AND io.approval_request_numeric_id IS NOT NULL
                  AND NOT (
                        COALESCE(io.approval_status::text, '') ILIKE '%cancel%'
                     OR COALESCE(io.normalized_status::text, '') ILIKE '%cancel%'
                     OR COALESCE(io.approval_status::text, '') ILIKE '%reject%'
                     OR COALESCE(io.normalized_status::text, '') ILIKE '%reject%'
                  )
            )
            SELECT DISTINCT internal_order_rows.internal_order_number
            FROM vw_sale_order_internal_order_bridge bridge
            JOIN internal_order_rows
                ON internal_order_rows.internal_order_id = bridge.internal_order_id
            WHERE bridge.so_id = :sales_order_id
            ORDER BY internal_order_rows.internal_order_number
        """)

        sales_order_summary_sql = text("""
            WITH summary_rows AS (
                SELECT *
                FROM vw_internal_order_rekap_summary
                WHERE internal_order_number IN :internal_order_numbers
            ), line_rows AS (
                SELECT *
                FROM vw_internal_order_rekap_lines
                WHERE internal_order_number IN :internal_order_numbers
            ), summary_agg AS (
                SELECT
                    STRING_AGG(DISTINCT internal_order_number, ', ' ORDER BY internal_order_number) AS internal_order_number,
                    STRING_AGG(DISTINCT company_name, ', ' ORDER BY company_name) FILTER (WHERE NULLIF(BTRIM(COALESCE(company_name::text, '')), '') IS NOT NULL) AS company_name,
                    SUM(io_reference_line_count) AS io_reference_line_count,
                    SUM(io_reference_amount) AS io_reference_amount,
                    SUM(rkb_kontribusi) AS rkb_kontribusi,
                    CASE
                        WHEN SUM(io_reference_amount) IS NULL OR SUM(io_reference_amount) = 0 THEN NULL
                        ELSE SUM(rkb_kontribusi) / NULLIF(SUM(io_reference_amount), 0)
                    END AS rkb_kontribusi_pct
                FROM summary_rows
            ), line_agg AS (
                SELECT
                    COUNT(*) AS product_count,
                    COUNT(*) FILTER (WHERE COALESCE(rkb_actual_qty, 0) <> 0 OR COALESCE(rkb_actual_subtotal, 0) <> 0) AS rkb_actual_product_count,
                    COUNT(*) FILTER (WHERE COALESCE(rop_qty, 0) <> 0 OR COALESCE(rop_subtotal, 0) <> 0) AS rop_product_count,
                    COUNT(*) FILTER (WHERE COALESCE(po_qty, 0) <> 0 OR COALESCE(po_subtotal, 0) <> 0) AS po_product_count,
                    SUM(rkb_actual_subtotal) AS rkb_actual_amount,
                    SUM(rkb_actual_subtotal) FILTER (WHERE product_trackability_class = 'TRACKABLE_PRODUCT') AS rkb_actual_trackable_amount,
                    SUM(rkb_actual_subtotal) FILTER (WHERE product_trackability_class <> 'TRACKABLE_PRODUCT') AS rkb_actual_non_trackable_amount,
                    SUM(rkb_actual_subtotal) FILTER (WHERE product_trackability_class = 'UNKNOWN_PRODUCT_CLASS') AS rkb_actual_unknown_class_amount,
                    SUM(rop_subtotal) AS rop_amount,
                    SUM(rop_subtotal) FILTER (WHERE product_trackability_class = 'TRACKABLE_PRODUCT') AS rop_trackable_amount,
                    SUM(rop_subtotal) FILTER (WHERE product_trackability_class <> 'TRACKABLE_PRODUCT') AS rop_non_trackable_amount,
                    SUM(rop_subtotal) FILTER (WHERE product_trackability_class = 'UNKNOWN_PRODUCT_CLASS') AS rop_unknown_class_amount,
                    SUM(po_subtotal) AS po_amount,
                    SUM(po_subtotal) FILTER (WHERE product_trackability_class = 'TRACKABLE_PRODUCT') AS po_trackable_amount,
                    SUM(po_subtotal) FILTER (WHERE product_trackability_class <> 'TRACKABLE_PRODUCT') AS po_non_trackable_amount,
                    SUM(po_subtotal) FILTER (WHERE product_trackability_class = 'UNKNOWN_PRODUCT_CLASS') AS po_unknown_class_amount,
                    SUM(not_yet_rop_amount) AS not_yet_rop_amount,
                    SUM(CASE WHEN COALESCE(rop_subtotal, 0) > 0 THEN GREATEST(COALESCE(rop_subtotal, 0) - COALESCE(rkb_actual_subtotal, 0), 0) ELSE 0 END) AS excess_rop_amount,
                    SUM(CASE WHEN COALESCE(po_subtotal, 0) > 0 THEN GREATEST(COALESCE(po_subtotal, 0) - COALESCE(rop_subtotal, 0), 0) ELSE 0 END) AS po_excess_amount,
                    SUM(po_received_qty) AS po_received_qty,
                    SUM(po_invoiced_qty) AS po_invoiced_qty,
                    COUNT(*) FILTER (WHERE mixed_uom_flag) AS mixed_uom_count,
                    COUNT(*) FILTER (WHERE product_presence_status = 'RKB_ONLY') AS rkb_only_count,
                    COUNT(*) FILTER (WHERE product_presence_status = 'ROP_ONLY') AS rop_only_count,
                    COUNT(*) FILTER (WHERE product_presence_status = 'PO_ONLY') AS po_only_count,
                    COUNT(*) FILTER (WHERE product_presence_status = 'RKB_ROP_PO') AS rkb_rop_po_count,
                    COUNT(*) FILTER (WHERE po_without_rop_flag) AS po_without_rop_count,
                    COUNT(*) FILTER (WHERE rop_without_po_flag) AS rop_without_po_count,
                    CASE WHEN SUM(po_qty) = 0 THEN NULL ELSE SUM(po_received_qty) / NULLIF(SUM(po_qty), 0) END AS received_ratio,
                    CASE WHEN SUM(po_qty) = 0 THEN NULL ELSE SUM(po_invoiced_qty) / NULLIF(SUM(po_qty), 0) END AS invoiced_ratio,
                    CASE WHEN SUM(rkb_actual_subtotal) = 0 THEN NULL ELSE SUM(rop_subtotal) / NULLIF(SUM(rkb_actual_subtotal), 0) END AS rop_progress_ratio,
                    CASE WHEN SUM(rkb_actual_subtotal) = 0 THEN NULL ELSE SUM(not_yet_rop_amount) / NULLIF(SUM(rkb_actual_subtotal), 0) END AS not_yet_rop_ratio
                FROM line_rows
            )
            SELECT
                summary_agg.internal_order_number,
                summary_agg.company_name,
                1 AS linked_sales_order_count,
                CAST(:sales_order_number AS text) AS linked_sales_order_numbers,
                TRUE AS has_sales_order_link,
                summary_agg.io_reference_line_count,
                summary_agg.io_reference_amount,
                line_agg.product_count,
                line_agg.rkb_actual_product_count,
                line_agg.rop_product_count,
                line_agg.po_product_count,
                line_agg.rkb_actual_amount,
                summary_agg.rkb_kontribusi,
                summary_agg.rkb_kontribusi_pct,
                line_agg.rkb_actual_trackable_amount,
                line_agg.rkb_actual_non_trackable_amount,
                line_agg.rkb_actual_unknown_class_amount,
                line_agg.rop_amount,
                line_agg.rop_trackable_amount,
                line_agg.rop_non_trackable_amount,
                line_agg.rop_unknown_class_amount,
                line_agg.po_amount,
                line_agg.po_trackable_amount,
                line_agg.po_non_trackable_amount,
                line_agg.po_unknown_class_amount,
                line_agg.not_yet_rop_amount,
                line_agg.excess_rop_amount,
                line_agg.po_excess_amount,
                line_agg.po_received_qty,
                line_agg.po_invoiced_qty,
                line_agg.mixed_uom_count,
                line_agg.rkb_only_count,
                line_agg.rop_only_count,
                line_agg.po_only_count,
                line_agg.rkb_rop_po_count,
                line_agg.po_without_rop_count,
                line_agg.rop_without_po_count,
                line_agg.received_ratio,
                line_agg.invoiced_ratio,
                line_agg.rop_progress_ratio,
                line_agg.not_yet_rop_ratio,
                'Sales Order linked IO context'::text AS comparison_basis,
                'IO-level material chain context'::text AS summary_scope
            FROM summary_agg
            CROSS JOIN line_agg
        """).bindparams(bindparam("internal_order_numbers", expanding=True))

        sales_order_breakdown_trackability_sql = text("""
            SELECT
                product_trackability_class,
                product_classification_reason,
                is_trackable_product,
                COUNT(*) AS product_count,
                SUM(rkb_actual_subtotal) AS rkb_actual_amount,
                SUM(rop_subtotal) AS rop_amount,
                SUM(po_subtotal) AS po_amount
            FROM vw_internal_order_rekap_lines
            WHERE internal_order_number IN :internal_order_numbers
            GROUP BY
                product_trackability_class,
                product_classification_reason,
                is_trackable_product
            ORDER BY
                product_trackability_class,
                product_classification_reason,
                is_trackable_product
        """).bindparams(bindparam("internal_order_numbers", expanding=True))

        sales_order_breakdown_presence_sql = text("""
            SELECT
                product_presence_status,
                COUNT(*) AS product_count,
                SUM(rkb_actual_subtotal) AS rkb_actual_amount,
                SUM(rop_subtotal) AS rop_amount,
                SUM(po_subtotal) AS po_amount
            FROM vw_internal_order_rekap_lines
            WHERE internal_order_number IN :internal_order_numbers
            GROUP BY product_presence_status
            ORDER BY product_presence_status
        """).bindparams(bindparam("internal_order_numbers", expanding=True))

        sales_order_lines_sql = text("""
            SELECT
                'LINKED_INTERNAL_ORDER'::text AS material_chain_source,
                lines.internal_order_number,
                lines.company_name,
                lines.linked_sales_order_count,
                lines.linked_sales_order_numbers,
                lines.has_sales_order_link,
                lines.product_key,
                lines.product_name,
                lines.product_trackability_class,
                lines.product_classification_reason,
                lines.is_trackable_product,
                lines.product_presence_status,
                lines.uom_summary,
                lines.rkb_actual_uom_summary,
                lines.rop_uom_summary,
                lines.po_uom_summary,
                lines.mixed_uom_flag,
                lines.rkb_actual_qty,
                lines.rkb_actual_unit_price,
                lines.rkb_actual_subtotal,
                lines.rkb_actual_request_summary,
                lines.rkb_actual_request_numeric_summary,
                lines.rop_qty,
                lines.rop_unit_price,
                lines.rop_subtotal,
                lines.rop_request_summary,
                lines.rop_request_numeric_summary,
                lines.po_qty,
                lines.po_unit_price,
                lines.po_subtotal,
                lines.po_received_qty,
                lines.po_invoiced_qty,
                po_refs.po_order_reference_summary,
                lines.not_yet_rop_qty,
                lines.not_yet_rop_amount,
                lines.excess_rop_qty,
                lines.excess_rop_amount,
                lines.po_without_rop_flag,
                lines.rop_without_po_flag,
                lines.comparison_scope
            FROM vw_internal_order_rekap_lines lines
            LEFT JOIN vw_internal_order_po_agg po_refs
                ON po_refs.internal_order_number = lines.internal_order_number
               AND po_refs.product_key = lines.product_key
            WHERE lines.internal_order_number IN :internal_order_numbers
            ORDER BY
                lines.internal_order_number,
                lines.product_trackability_class,
                lines.product_presence_status,
                COALESCE(lines.rkb_actual_subtotal, 0) DESC,
                lines.product_key
        """).bindparams(bindparam("internal_order_numbers", expanding=True))


        direct_lines_sql = text(r"""
            WITH rkb_rows AS (
                SELECT
                    COALESCE(NULLIF(BTRIM((regexp_match(COALESCE(apl.product_name::text, ''), '^\[([^\]]+)\]'))[1]), ''), NULLIF(LOWER(REGEXP_REPLACE(BTRIM(COALESCE(apl.product_name::text, '')), '[[:space:]]+', ' ', 'g')), '')) AS product_key,
                    NULLIF(BTRIM(COALESCE(apl.product_name::text, '')), '') AS product_name,
                    NULLIF(BTRIM(COALESCE(apl.unit_of_measure::text, '')), '') AS unit_of_measure,
                    COALESCE(apl.planned_quantity, 0)::numeric AS quantity,
                    COALESCE(apl.planned_unit_price, 0)::numeric AS unit_price,
                    COALESCE(apl.planned_subtotal, 0)::numeric AS subtotal,
                    NULLIF(BTRIM(COALESCE(apl.approval_request_id::text, '')), '') AS request_reference,
                    apl.approval_request_numeric_id::text AS request_numeric_reference,
                    NULLIF(BTRIM(COALESCE(apl.company_name::text, '')), '') AS company_name
                FROM vw_approval_product_line_context apl
                WHERE apl.normalized_jo_number = :sales_order_number
                  AND apl.approval_business_type = 'RKB_PLANNING'
                  AND apl.is_valid_for_metrics
                  AND NOT (COALESCE(apl.approval_status::text, '') ILIKE '%cancel%' OR COALESCE(apl.normalized_status::text, '') ILIKE '%cancel%' OR COALESCE(apl.approval_status::text, '') ILIKE '%reject%' OR COALESCE(apl.normalized_status::text, '') ILIKE '%reject%')
            ), rop_rows AS (
                SELECT
                    COALESCE(NULLIF(BTRIM((regexp_match(COALESCE(apl.product_name::text, ''), '^\[([^\]]+)\]'))[1]), ''), NULLIF(LOWER(REGEXP_REPLACE(BTRIM(COALESCE(apl.product_name::text, '')), '[[:space:]]+', ' ', 'g')), '')) AS product_key,
                    NULLIF(BTRIM(COALESCE(apl.product_name::text, '')), '') AS product_name,
                    NULLIF(BTRIM(COALESCE(apl.unit_of_measure::text, '')), '') AS unit_of_measure,
                    COALESCE(apl.planned_quantity, 0)::numeric AS quantity,
                    COALESCE(apl.planned_unit_price, 0)::numeric AS unit_price,
                    COALESCE(apl.planned_subtotal, 0)::numeric AS subtotal,
                    NULLIF(BTRIM(COALESCE(apl.approval_request_id::text, '')), '') AS request_reference,
                    apl.approval_request_numeric_id::text AS request_numeric_reference,
                    NULLIF(BTRIM(COALESCE(apl.company_name::text, '')), '') AS company_name
                FROM vw_approval_product_line_context apl
                WHERE apl.normalized_jo_number = :sales_order_number
                  AND apl.approval_business_type = 'ROP_PROCUREMENT_REQUEST'
                  AND apl.is_valid_for_metrics
                  AND NOT (COALESCE(apl.approval_status::text, '') ILIKE '%cancel%' OR COALESCE(apl.normalized_status::text, '') ILIKE '%cancel%' OR COALESCE(apl.approval_status::text, '') ILIKE '%reject%' OR COALESCE(apl.normalized_status::text, '') ILIKE '%reject%')
            ), po_rows AS (
                SELECT
                    COALESCE(NULLIF(BTRIM((regexp_match(COALESCE(po.product_name::text, ''), '^\[([^\]]+)\]'))[1]), ''), NULLIF(LOWER(REGEXP_REPLACE(BTRIM(COALESCE(po.product_name::text, '')), '[[:space:]]+', ' ', 'g')), '')) AS product_key,
                    NULLIF(BTRIM(COALESCE(po.product_name::text, '')), '') AS product_name,
                    NULLIF(BTRIM(COALESCE(po.unit_of_measure::text, '')), '') AS unit_of_measure,
                    COALESCE(po.ordered_quantity, 0)::numeric AS quantity,
                    COALESCE(po.received_quantity, 0)::numeric AS received_quantity,
                    COALESCE(po.invoiced_quantity, 0)::numeric AS invoiced_quantity,
                    COALESCE(po.unit_price, 0)::numeric AS unit_price,
                    COALESCE(po.line_subtotal, 0)::numeric AS subtotal,
                    NULLIF(BTRIM(COALESCE(po.purchase_order_reference::text, '')), '') AS purchase_order_reference,
                    NULLIF(BTRIM(COALESCE(po.company_name::text, '')), '') AS company_name
                FROM vw_procurement_lines po
                WHERE po.normalized_jo_number = :sales_order_number
                  AND po.is_valid_for_metrics
                  AND NOT (COALESCE(po.purchase_line_state::text, '') ILIKE '%cancel%' OR COALESCE(po.normalized_status::text, '') ILIKE '%cancel%' OR COALESCE(po.purchase_line_state::text, '') ILIKE '%reject%' OR COALESCE(po.normalized_status::text, '') ILIKE '%reject%')
            ), stock_rows AS (
                SELECT
                    COALESCE(NULLIF(BTRIM((regexp_match(COALESCE(sml.product_id::text, ''), '^\[([^\]]+)\]'))[1]), ''), NULLIF(LOWER(REGEXP_REPLACE(BTRIM(COALESCE(sml.product_id::text, '')), '[[:space:]]+', ' ', 'g')), '')) AS product_key,
                    NULLIF(BTRIM(COALESCE(sml.product_id::text, '')), '') AS product_name,
                    NULLIF(BTRIM(COALESCE(sml.product_uom_id::text, '')), '') AS unit_of_measure,
                    COALESCE(sml.quantity, 0)::numeric AS quantity,
                    NULLIF(BTRIM(COALESCE(sml.company_id::text, '')), '') AS company_name
                FROM stock_move_line sml
                WHERE BTRIM(COALESCE(sml.x_studio_source_document::text, '')) = :sales_order_number
                  AND COALESCE(sml.state::text, '') NOT ILIKE '%cancel%'
            ), rkb AS (
                SELECT product_key, MAX(product_name) product_name, STRING_AGG(DISTINCT unit_of_measure, ' ; ' ORDER BY unit_of_measure) FILTER (WHERE unit_of_measure IS NOT NULL) uom_summary, SUM(quantity) qty, CASE WHEN SUM(quantity)=0 THEN NULL ELSE SUM(subtotal)/NULLIF(SUM(quantity),0) END unit_price, SUM(subtotal) subtotal, STRING_AGG(DISTINCT request_reference, ', ' ORDER BY request_reference) FILTER (WHERE request_reference IS NOT NULL) request_summary, STRING_AGG(DISTINCT request_numeric_reference, ', ' ORDER BY request_numeric_reference) FILTER (WHERE request_numeric_reference IS NOT NULL) request_numeric_summary, MAX(company_name) company_name FROM rkb_rows WHERE product_key IS NOT NULL GROUP BY product_key
            ), rop AS (
                SELECT product_key, MAX(product_name) product_name, STRING_AGG(DISTINCT unit_of_measure, ' ; ' ORDER BY unit_of_measure) FILTER (WHERE unit_of_measure IS NOT NULL) uom_summary, SUM(quantity) qty, CASE WHEN SUM(quantity)=0 THEN NULL ELSE SUM(subtotal)/NULLIF(SUM(quantity),0) END unit_price, SUM(subtotal) subtotal, STRING_AGG(DISTINCT request_reference, ', ' ORDER BY request_reference) FILTER (WHERE request_reference IS NOT NULL) request_summary, STRING_AGG(DISTINCT request_numeric_reference, ', ' ORDER BY request_numeric_reference) FILTER (WHERE request_numeric_reference IS NOT NULL) request_numeric_summary, MAX(company_name) company_name FROM rop_rows WHERE product_key IS NOT NULL GROUP BY product_key
            ), po AS (
                SELECT product_key, MAX(product_name) product_name, STRING_AGG(DISTINCT unit_of_measure, ' ; ' ORDER BY unit_of_measure) FILTER (WHERE unit_of_measure IS NOT NULL) uom_summary, SUM(quantity) qty, SUM(received_quantity) received_qty, SUM(invoiced_quantity) invoiced_qty, CASE WHEN SUM(quantity)=0 THEN NULL ELSE SUM(subtotal)/NULLIF(SUM(quantity),0) END unit_price, SUM(subtotal) subtotal, STRING_AGG(DISTINCT purchase_order_reference, ', ' ORDER BY purchase_order_reference) FILTER (WHERE purchase_order_reference IS NOT NULL) po_refs, MAX(company_name) company_name FROM po_rows WHERE product_key IS NOT NULL GROUP BY product_key
            ), stock AS (
                SELECT product_key, MAX(product_name) product_name, STRING_AGG(DISTINCT unit_of_measure, ' ; ' ORDER BY unit_of_measure) FILTER (WHERE unit_of_measure IS NOT NULL) uom_summary, SUM(quantity) qty, MAX(company_name) company_name FROM stock_rows WHERE product_key IS NOT NULL GROUP BY product_key
            ), universe AS (
                SELECT product_key FROM rkb UNION SELECT product_key FROM rop UNION SELECT product_key FROM po UNION SELECT product_key FROM stock
            ), line_rows AS (
                SELECT universe.product_key, COALESCE(rkb.product_name, rop.product_name, po.product_name, stock.product_name) product_name, COALESCE(rkb.company_name, rop.company_name, po.company_name, stock.company_name) company_name, rkb.uom_summary rkb_uom, rop.uom_summary rop_uom, po.uom_summary po_uom, stock.uom_summary stock_uom, COALESCE(rkb.qty,0)::numeric rkb_qty, rkb.unit_price rkb_unit_price, COALESCE(rkb.subtotal,0)::numeric rkb_subtotal, rkb.request_summary rkb_request_summary, rkb.request_numeric_summary rkb_request_numeric_summary, COALESCE(rop.qty,0)::numeric rop_qty, rop.unit_price rop_unit_price, COALESCE(rop.subtotal,0)::numeric rop_subtotal, rop.request_summary rop_request_summary, rop.request_numeric_summary rop_request_numeric_summary, COALESCE(po.qty,0)::numeric po_qty, po.unit_price po_unit_price, COALESCE(po.subtotal,0)::numeric po_subtotal, COALESCE(po.received_qty,0)::numeric po_received_qty, COALESCE(po.invoiced_qty,0)::numeric po_invoiced_qty, po.po_refs po_order_reference_summary, COALESCE(stock.qty,0)::numeric stock_qty, CONCAT_WS(' ; ', rkb.uom_summary, rop.uom_summary, po.uom_summary, stock.uom_summary) uom_summary FROM universe LEFT JOIN rkb ON rkb.product_key=universe.product_key LEFT JOIN rop ON rop.product_key=universe.product_key LEFT JOIN po ON po.product_key=universe.product_key LEFT JOIN stock ON stock.product_key=universe.product_key
            )
            SELECT
                CASE WHEN rkb_qty=0 AND rop_qty=0 AND po_qty=0 AND stock_qty<>0 THEN 'FROM_STOCK' ELSE 'DIRECT_SALES_ORDER' END AS material_chain_source,
                NULL::text AS internal_order_number,
                company_name,
                1 AS linked_sales_order_count,
                CAST(:sales_order_number AS text) AS linked_sales_order_numbers,
                TRUE AS has_sales_order_link,
                product_key,
                product_name,
                CASE WHEN COALESCE(product_key,'') ILIKE '!!%%' OR COALESCE(product_name,'') ILIKE '!!%%' THEN 'NON_TRACKABLE_OTHERS' WHEN COALESCE(product_key,'') ILIKE '%%others%%' OR COALESCE(product_name,'') ILIKE '%%others%%' THEN 'NON_TRACKABLE_OTHERS' WHEN COALESCE(product_key,'') ILIKE '%%sisa budget%%' OR COALESCE(product_name,'') ILIKE '%%sisa budget%%' OR COALESCE(product_key,'') ILIKE '%%estimator%%' OR COALESCE(product_name,'') ILIKE '%%estimator%%' OR COALESCE(product_key,'') ILIKE '%%jasa%%' OR COALESCE(product_name,'') ILIKE '%%jasa%%' OR COALESCE(product_key,'') ILIKE '%%machining%%' OR COALESCE(product_name,'') ILIKE '%%machining%%' THEN 'BUDGET_SERVICE_ADJUSTMENT' WHEN COALESCE(product_key,'') ~ '^\[[0-9]{5}\]' OR COALESCE(product_name,'') ~ '^\[[0-9]{5}\]' THEN 'TRACKABLE_PRODUCT' ELSE 'UNKNOWN_PRODUCT_CLASS' END AS product_trackability_class,
                CASE WHEN COALESCE(product_key,'') ILIKE '!!%%' OR COALESCE(product_name,'') ILIKE '!!%%' THEN 'DOUBLE_BANG_OTHERS' WHEN COALESCE(product_key,'') ILIKE '%%others%%' OR COALESCE(product_name,'') ILIKE '%%others%%' THEN 'CONTAINS_OTHERS' WHEN COALESCE(product_key,'') ILIKE '%%sisa budget%%' OR COALESCE(product_name,'') ILIKE '%%sisa budget%%' OR COALESCE(product_key,'') ILIKE '%%estimator%%' OR COALESCE(product_name,'') ILIKE '%%estimator%%' OR COALESCE(product_key,'') ILIKE '%%jasa%%' OR COALESCE(product_name,'') ILIKE '%%jasa%%' OR COALESCE(product_key,'') ILIKE '%%machining%%' OR COALESCE(product_name,'') ILIKE '%%machining%%' THEN 'BUDGET_SERVICE_TEXT' WHEN COALESCE(product_key,'') ~ '^\[[0-9]{5}\]' OR COALESCE(product_name,'') ~ '^\[[0-9]{5}\]' THEN 'BRACKETED_PRODUCT_CODE' ELSE 'UNKNOWN_FALLBACK' END AS product_classification_reason,
                (COALESCE(product_key,'') ~ '^\[[0-9]{5}\]' OR COALESCE(product_name,'') ~ '^\[[0-9]{5}\]') AS is_trackable_product,
                CASE WHEN rkb_qty<>0 AND rop_qty<>0 AND po_qty<>0 THEN 'RKB_ROP_PO' WHEN rkb_qty<>0 AND rop_qty<>0 AND po_qty=0 THEN 'RKB_ROP' WHEN rkb_qty<>0 AND rop_qty=0 AND po_qty<>0 THEN 'RKB_PO' WHEN rkb_qty=0 AND rop_qty<>0 AND po_qty<>0 THEN 'ROP_PO' WHEN rkb_qty<>0 AND rop_qty=0 AND po_qty=0 THEN 'RKB_ONLY' WHEN rkb_qty=0 AND rop_qty<>0 AND po_qty=0 THEN 'ROP_ONLY' WHEN rkb_qty=0 AND rop_qty=0 AND po_qty<>0 THEN 'PO_ONLY' ELSE 'UNKNOWN_PRODUCT_CLASS' END AS product_presence_status,
                uom_summary,
                rkb_uom AS rkb_actual_uom_summary,
                rop_uom AS rop_uom_summary,
                po_uom AS po_uom_summary,
                (SELECT COUNT(DISTINCT NULLIF(BTRIM(uom_value), '')) FROM unnest(ARRAY[rkb_uom, rop_uom, po_uom]) AS uom_value) > 1 AS mixed_uom_flag,
                rkb_qty AS rkb_actual_qty,
                rkb_unit_price AS rkb_actual_unit_price,
                rkb_subtotal AS rkb_actual_subtotal,
                rkb_request_summary AS rkb_actual_request_summary,
                rkb_request_numeric_summary AS rkb_actual_request_numeric_summary,
                rop_qty,
                rop_unit_price,
                rop_subtotal,
                rop_request_summary,
                rop_request_numeric_summary,
                po_qty,
                po_unit_price,
                po_subtotal,
                po_received_qty,
                po_invoiced_qty,
                po_order_reference_summary,
                0::numeric AS not_yet_rop_qty,
                0::numeric AS not_yet_rop_amount,
                GREATEST(rop_qty-rkb_qty,0)::numeric AS excess_rop_qty,
                GREATEST(rop_subtotal-rkb_subtotal,0)::numeric AS excess_rop_amount,
                GREATEST(po_subtotal-rop_subtotal,0)::numeric AS po_excess_amount,
                (po_qty<>0 AND rop_qty=0) AS po_without_rop_flag,
                (rop_qty<>0 AND po_qty=0) AS rop_without_po_flag,
                'DIRECT_SO_JO_CONTEXT'::text AS comparison_scope
            FROM line_rows
            ORDER BY product_trackability_class, product_presence_status, COALESCE(rkb_subtotal, 0) DESC, product_key
        """)


        def _aggregate_sales_order_lines(
            lines: list[dict],
            selected_so_number: str,
            linked_internal_orders: list[str],
            sales_order_amount: Decimal,
            linked_context: dict | None = None,
        ) -> dict:
            linked_context = linked_context or {}
            rkb_amount = sum((_decimal_value(row.get("rkb_actual_subtotal")) for row in lines), Decimal("0"))
            rop_amount = sum((_decimal_value(row.get("rop_subtotal")) for row in lines), Decimal("0"))
            po_amount = sum((_decimal_value(row.get("po_subtotal")) for row in lines), Decimal("0"))
            po_qty = sum((_decimal_value(row.get("po_qty")) for row in lines), Decimal("0"))
            po_received_qty = sum((_decimal_value(row.get("po_received_qty")) for row in lines), Decimal("0"))
            po_invoiced_qty = sum((_decimal_value(row.get("po_invoiced_qty")) for row in lines), Decimal("0"))
            rop_excess_amount = sum((
                max(_decimal_value(row.get("rop_subtotal")) - _decimal_value(row.get("rkb_actual_subtotal")), Decimal("0"))
                if _decimal_value(row.get("rop_subtotal")) > 0 else Decimal("0")
                for row in lines
            ), Decimal("0"))
            po_excess_amount = sum((
                max(_decimal_value(row.get("po_subtotal")) - _decimal_value(row.get("rop_subtotal")), Decimal("0"))
                if _decimal_value(row.get("po_subtotal")) > 0 else Decimal("0")
                for row in lines
            ), Decimal("0"))
            so_rkb_kontribusi = sales_order_amount - rkb_amount if sales_order_amount != 0 else None
            so_rkb_kontribusi_pct = _ratio_value(so_rkb_kontribusi, sales_order_amount) if so_rkb_kontribusi is not None else None

            def amount_for(class_name: str, field_name: str) -> Decimal:
                return sum((_decimal_value(row.get(field_name)) for row in lines if row.get("product_trackability_class") == class_name), Decimal("0"))

            return {
                "internal_order_number": ", ".join(linked_internal_orders) if linked_internal_orders else None,
                "company_name": ", ".join(sorted({str(row.get("company_name")) for row in lines if row.get("company_name")})) or None,
                "linked_sales_order_count": 1,
                "linked_sales_order_numbers": selected_so_number,
                "has_sales_order_link": True,
                "sales_order_amount": sales_order_amount,
                "so_rkb_kontribusi": so_rkb_kontribusi,
                "so_rkb_kontribusi_pct": so_rkb_kontribusi_pct,
                "io_reference_line_count": linked_context.get("io_reference_line_count"),
                "io_reference_amount": linked_context.get("io_reference_amount"),
                "product_count": len(lines),
                "rkb_actual_product_count": sum(1 for row in lines if _decimal_value(row.get("rkb_actual_qty")) != 0 or _decimal_value(row.get("rkb_actual_subtotal")) != 0),
                "rop_product_count": sum(1 for row in lines if _decimal_value(row.get("rop_qty")) != 0 or _decimal_value(row.get("rop_subtotal")) != 0),
                "po_product_count": sum(1 for row in lines if _decimal_value(row.get("po_qty")) != 0 or _decimal_value(row.get("po_subtotal")) != 0),
                "rkb_actual_amount": rkb_amount,
                "rkb_kontribusi": linked_context.get("rkb_kontribusi"),
                "rkb_kontribusi_pct": linked_context.get("rkb_kontribusi_pct"),
                "rkb_actual_trackable_amount": amount_for("TRACKABLE_PRODUCT", "rkb_actual_subtotal"),
                "rkb_actual_non_trackable_amount": sum((_decimal_value(row.get("rkb_actual_subtotal")) for row in lines if row.get("product_trackability_class") != "TRACKABLE_PRODUCT"), Decimal("0")),
                "rkb_actual_unknown_class_amount": amount_for("UNKNOWN_PRODUCT_CLASS", "rkb_actual_subtotal"),
                "rop_amount": rop_amount,
                "rop_trackable_amount": amount_for("TRACKABLE_PRODUCT", "rop_subtotal"),
                "rop_non_trackable_amount": sum((_decimal_value(row.get("rop_subtotal")) for row in lines if row.get("product_trackability_class") != "TRACKABLE_PRODUCT"), Decimal("0")),
                "rop_unknown_class_amount": amount_for("UNKNOWN_PRODUCT_CLASS", "rop_subtotal"),
                "po_amount": po_amount,
                "po_trackable_amount": amount_for("TRACKABLE_PRODUCT", "po_subtotal"),
                "po_non_trackable_amount": sum((_decimal_value(row.get("po_subtotal")) for row in lines if row.get("product_trackability_class") != "TRACKABLE_PRODUCT"), Decimal("0")),
                "po_unknown_class_amount": amount_for("UNKNOWN_PRODUCT_CLASS", "po_subtotal"),
                "not_yet_rop_amount": sum((_decimal_value(row.get("not_yet_rop_amount")) for row in lines), Decimal("0")),
                "excess_rop_amount": rop_excess_amount,
                "po_excess_amount": po_excess_amount,
                "po_received_qty": po_received_qty,
                "po_invoiced_qty": po_invoiced_qty,
                "mixed_uom_count": sum(1 for row in lines if row.get("mixed_uom_flag")),
                "rkb_only_count": sum(1 for row in lines if row.get("product_presence_status") == "RKB_ONLY"),
                "rop_only_count": sum(1 for row in lines if row.get("product_presence_status") == "ROP_ONLY"),
                "po_only_count": sum(1 for row in lines if row.get("product_presence_status") == "PO_ONLY"),
                "rkb_rop_po_count": sum(1 for row in lines if row.get("product_presence_status") == "RKB_ROP_PO"),
                "po_without_rop_count": sum(1 for row in lines if row.get("po_without_rop_flag")),
                "rop_without_po_count": sum(1 for row in lines if row.get("rop_without_po_flag")),
                "received_ratio": _ratio_value(po_received_qty, po_qty),
                "invoiced_ratio": _ratio_value(po_invoiced_qty, po_qty),
                "rop_progress_ratio": _ratio_value(rop_amount, rkb_amount),
                "not_yet_rop_ratio": _ratio_value(sum((_decimal_value(row.get("not_yet_rop_amount")) for row in lines), Decimal("0")), rkb_amount),
                "comparison_basis": "Sales Order linked IO and/or direct SO/JO context",
                "summary_scope": "Traceability context only; not product allocation or accounting margin",
            }

        def _aggregate_sales_order_breakdowns(lines: list[dict]) -> tuple[list[dict], list[dict]]:
            trackability_groups = {}
            presence_groups = {}
            for row in lines:
                t_key = (row.get("product_trackability_class"), row.get("product_classification_reason"), bool(row.get("is_trackable_product")))
                t_group = trackability_groups.setdefault(t_key, {"product_trackability_class": t_key[0], "product_classification_reason": t_key[1], "is_trackable_product": t_key[2], "product_count": 0, "rkb_actual_amount": Decimal("0"), "rop_amount": Decimal("0"), "po_amount": Decimal("0")})
                t_group["product_count"] += 1
                t_group["rkb_actual_amount"] += _decimal_value(row.get("rkb_actual_subtotal"))
                t_group["rop_amount"] += _decimal_value(row.get("rop_subtotal"))
                t_group["po_amount"] += _decimal_value(row.get("po_subtotal"))
                p_key = row.get("product_presence_status") or "UNKNOWN_PRODUCT_CLASS"
                p_group = presence_groups.setdefault(p_key, {"product_presence_status": p_key, "product_count": 0, "rkb_actual_amount": Decimal("0"), "rop_amount": Decimal("0"), "po_amount": Decimal("0")})
                p_group["product_count"] += 1
                p_group["rkb_actual_amount"] += _decimal_value(row.get("rkb_actual_subtotal"))
                p_group["rop_amount"] += _decimal_value(row.get("rop_subtotal"))
                p_group["po_amount"] += _decimal_value(row.get("po_subtotal"))
            return (
                [{key: _json_safe(value) for key, value in group.items()} for group in sorted(trackability_groups.values(), key=lambda item: (str(item["product_trackability_class"]), str(item["product_classification_reason"]), str(item["is_trackable_product"])))],
                [{key: _json_safe(value) for key, value in group.items()} for group in sorted(presence_groups.values(), key=lambda item: str(item["product_presence_status"]))],
            )

        pg = PostgresClient()
        try:
            with pg.engine.connect() as conn:
                selected_sales_order = conn.execute(
                    selected_sales_order_sql,
                    {"sales_order_number": normalized_sales_order_number},
                ).mappings().first()

                if selected_sales_order is None:
                    raise HTTPException(
                        status_code=404,
                        detail={
                            "error": "Sales Order not found or not active.",
                            "sales_order_number": normalized_sales_order_number,
                        },
                    )

                linked_internal_orders = [
                    row["internal_order_number"]
                    for row in conn.execute(
                        linked_internal_orders_sql,
                        {"sales_order_id": selected_sales_order["sales_order_id"]},
                    ).mappings().fetchall()
                    if row["internal_order_number"]
                ]

                linked_context = {}
                linked_lines = []
                if linked_internal_orders:
                    params = {
                        "internal_order_numbers": linked_internal_orders,
                        "sales_order_number": selected_sales_order["sales_order_number"],
                    }
                    summary_row = conn.execute(sales_order_summary_sql, params).fetchone()
                    linked_context = {
                        key: _json_safe(value)
                        for key, value in summary_row._mapping.items()
                    } if summary_row is not None else {}
                    linked_lines = [
                        {key: _json_safe(value) for key, value in row._mapping.items()}
                        for row in conn.execute(sales_order_lines_sql, params).fetchall()
                    ]

                direct_lines = [
                    {key: _json_safe(value) for key, value in row._mapping.items()}
                    for row in conn.execute(
                        direct_lines_sql,
                        {"sales_order_number": selected_sales_order["sales_order_number"]},
                    ).fetchall()
                ]

                lines = linked_lines + direct_lines
                summary = _aggregate_sales_order_lines(
                    lines,
                    selected_sales_order["sales_order_number"],
                    linked_internal_orders,
                    _decimal_value(selected_sales_order.get("sales_order_amount")),
                    linked_context,
                )
                summary = {key: _json_safe(value) for key, value in summary.items()}
                trackability_breakdowns, presence_breakdowns = _aggregate_sales_order_breakdowns(lines)

            linked_internal_order_numbers = ", ".join(linked_internal_orders)
            direct_sales_order_chain_count = sum(1 for row in lines if row.get("material_chain_source") == "DIRECT_SALES_ORDER")
            from_stock_chain_count = sum(1 for row in lines if row.get("material_chain_source") == "FROM_STOCK")
            empty_state_message = None
            if not lines:
                empty_state_message = "Sales Order found, but no linked IO or direct RKB / ROP / PO material chain was found."

            metadata = {
                "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
                "perspective": "sales_order",
                "comparison_basis": summary.get("comparison_basis"),
                "summary_scope": summary.get("summary_scope"),
                "line_count": len(lines),
                "selected_sales_order_id": _json_safe(selected_sales_order["sales_order_id"]),
                "selected_sales_order_number": selected_sales_order["sales_order_number"],
                "linked_internal_order_count": len(linked_internal_orders),
                "linked_internal_order_numbers": linked_internal_order_numbers,
                "direct_sales_order_chain_count": direct_sales_order_chain_count,
                "from_stock_chain_count": from_stock_chain_count,
                "has_linked_internal_order": bool(linked_internal_orders),
                "has_internal_order_link": bool(linked_internal_orders),
                "has_direct_sales_order_chain": direct_sales_order_chain_count > 0,
                "context_note": context_note,
                "empty_state_message": empty_state_message,
                "warnings": [empty_state_message] if empty_state_message else [],
            }

            return {
                "status": "success",
                "perspective": "sales_order",
                "sales_order_number": selected_sales_order["sales_order_number"],
                "internal_order_number": summary.get("internal_order_number"),
                "summary": summary,
                "breakdowns": {
                    "by_trackability_class": trackability_breakdowns,
                    "by_product_presence_status": presence_breakdowns,
                },
                "lines": lines,
                "metadata": metadata,
            }
        except HTTPException:
            raise
        except Exception as e:
            logger.error("Sales Order Perspective material tracking query failed", sales_order_number=normalized_sales_order_number, error=str(e))
            raise HTTPException(status_code=500, detail=str(e))
        finally:
            pg.close()

    if internal_order_number is None or not internal_order_number.strip():
        raise HTTPException(
            status_code=400,
            detail={"error": "internal_order_number query parameter is required."},
        )

    normalized_internal_order_number = internal_order_number.strip()

    summary_sql = text("""
        SELECT
            internal_order_number,
            company_name,
            linked_sales_order_count,
            linked_sales_order_numbers,
            has_sales_order_link,
            io_reference_line_count,
            io_reference_amount,
            product_count,
            rkb_actual_product_count,
            rop_product_count,
            po_product_count,
            rkb_actual_amount,
            rkb_kontribusi,
            rkb_kontribusi_pct,
            rkb_actual_trackable_amount,
            rkb_actual_non_trackable_amount,
            rkb_actual_unknown_class_amount,
            rop_amount,
            rop_trackable_amount,
            rop_non_trackable_amount,
            rop_unknown_class_amount,
            po_amount,
            po_trackable_amount,
            po_non_trackable_amount,
            po_unknown_class_amount,
            not_yet_rop_amount,
            excess_rop_amount,
            po_received_qty,
            po_invoiced_qty,
            mixed_uom_count,
            rkb_only_count,
            rop_only_count,
            po_only_count,
            rkb_rop_po_count,
            po_without_rop_count,
            rop_without_po_count,
            received_ratio,
            invoiced_ratio,
            rop_progress_ratio,
            not_yet_rop_ratio,
            comparison_basis,
            summary_scope
        FROM vw_internal_order_rekap_summary
        WHERE internal_order_number = :internal_order_number
    """)

    breakdown_trackability_sql = text("""
        SELECT
            product_trackability_class,
            product_classification_reason,
            is_trackable_product,
            COUNT(*) AS product_count,
            SUM(rkb_actual_subtotal) AS rkb_actual_amount,
            SUM(rop_subtotal) AS rop_amount,
            SUM(po_subtotal) AS po_amount
        FROM vw_internal_order_rekap_lines
        WHERE internal_order_number = :internal_order_number
        GROUP BY
            product_trackability_class,
            product_classification_reason,
            is_trackable_product
        ORDER BY
            product_trackability_class,
            product_classification_reason,
            is_trackable_product
    """)

    breakdown_presence_sql = text("""
        SELECT
            product_presence_status,
            COUNT(*) AS product_count,
            SUM(rkb_actual_subtotal) AS rkb_actual_amount,
            SUM(rop_subtotal) AS rop_amount,
            SUM(po_subtotal) AS po_amount
        FROM vw_internal_order_rekap_lines
        WHERE internal_order_number = :internal_order_number
        GROUP BY product_presence_status
        ORDER BY product_presence_status
    """)

    lines_sql = text("""
        SELECT
            'INTERNAL_ORDER'::text AS material_chain_source,
            lines.internal_order_number,
            lines.company_name,
            lines.linked_sales_order_count,
            lines.linked_sales_order_numbers,
            lines.has_sales_order_link,
            lines.product_key,
            lines.product_name,
            lines.product_trackability_class,
            lines.product_classification_reason,
            lines.is_trackable_product,
            lines.product_presence_status,
            lines.uom_summary,
            lines.rkb_actual_uom_summary,
            lines.rop_uom_summary,
            lines.po_uom_summary,
            lines.mixed_uom_flag,
            lines.rkb_actual_qty,
            lines.rkb_actual_unit_price,
            lines.rkb_actual_subtotal,
            lines.rkb_actual_request_summary,
            lines.rkb_actual_request_numeric_summary,
            lines.rop_qty,
            lines.rop_unit_price,
            lines.rop_subtotal,
            lines.rop_request_summary,
            lines.rop_request_numeric_summary,
            lines.po_qty,
            lines.po_unit_price,
            lines.po_subtotal,
            lines.po_received_qty,
            lines.po_invoiced_qty,
            po_refs.po_order_reference_summary,
            lines.not_yet_rop_qty,
            lines.not_yet_rop_amount,
            lines.excess_rop_qty,
            lines.excess_rop_amount,
            lines.po_without_rop_flag,
            lines.rop_without_po_flag,
            lines.comparison_scope
        FROM vw_internal_order_rekap_lines lines
        LEFT JOIN vw_internal_order_po_agg po_refs
            ON po_refs.internal_order_number = lines.internal_order_number
           AND po_refs.product_key = lines.product_key
        WHERE lines.internal_order_number = :internal_order_number
        ORDER BY
            lines.product_trackability_class,
            lines.product_presence_status,
            COALESCE(lines.rkb_actual_subtotal, 0) DESC,
            lines.product_key
    """)

    pg = PostgresClient()
    try:
        with pg.engine.connect() as conn:
            summary_result = conn.execute(
                summary_sql,
                {"internal_order_number": normalized_internal_order_number},
            ).fetchone()
            if summary_result is None:
                raise HTTPException(
                    status_code=404,
                    detail={
                        "error": "Internal Order not found.",
                        "internal_order_number": normalized_internal_order_number,
                    },
                )

            summary = {
                key: _json_safe(value)
                for key, value in summary_result._mapping.items()
            }

            trackability_breakdowns = [
                {key: _json_safe(value) for key, value in row._mapping.items()}
                for row in conn.execute(
                    breakdown_trackability_sql,
                    {"internal_order_number": normalized_internal_order_number},
                ).fetchall()
            ]

            presence_breakdowns = [
                {key: _json_safe(value) for key, value in row._mapping.items()}
                for row in conn.execute(
                    breakdown_presence_sql,
                    {"internal_order_number": normalized_internal_order_number},
                ).fetchall()
            ]

            lines = [
                {key: _json_safe(value) for key, value in row._mapping.items()}
                for row in conn.execute(
                    lines_sql,
                    {"internal_order_number": normalized_internal_order_number},
                ).fetchall()
            ]

            summary["excess_rop_amount"] = _json_safe(sum((
                max(_decimal_value(row.get("rop_subtotal")) - _decimal_value(row.get("rkb_actual_subtotal")), Decimal("0"))
                if _decimal_value(row.get("rop_subtotal")) > 0 else Decimal("0")
                for row in lines
            ), Decimal("0")))
            summary["po_excess_amount"] = _json_safe(sum((
                max(_decimal_value(row.get("po_subtotal")) - _decimal_value(row.get("rop_subtotal")), Decimal("0"))
                if _decimal_value(row.get("po_subtotal")) > 0 else Decimal("0")
                for row in lines
            ), Decimal("0")))

        metadata = {
            "generated_at": datetime.utcnow().replace(microsecond=0).isoformat() + "Z",
            "comparison_basis": summary.get("comparison_basis"),
            "summary_scope": summary.get("summary_scope"),
            "line_count": len(lines),
            "warnings": _build_internal_order_rekap_warnings(summary),
        }

        return {
            "status": "success",
            "internal_order_number": normalized_internal_order_number,
            "summary": summary,
            "breakdowns": {
                "by_trackability_class": trackability_breakdowns,
                "by_product_presence_status": presence_breakdowns,
            },
            "lines": lines,
            "metadata": metadata,
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Internal Order Rekap API query failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        pg.close()


@app.get("/health", response_model=HealthResponse, tags=["Health"])
async def health_check():
    """
    Check health status of all services.

    Returns connection status for Odoo and PostgreSQL,
    and whether the scheduler is running.
    """
    try:
        from src.clients.odoo_client import OdooClient

        odoo = OdooClient()
        pg = PostgresClient()

        odoo_connected = odoo.test_connection()
        postgres_connected = pg.test_connection()

        odoo.close()
        pg.close()

        status = "healthy" if (odoo_connected and postgres_connected) else "degraded"

        return HealthResponse(
            status=status,
            odoo_connected=odoo_connected,
            postgres_connected=postgres_connected,
            scheduler_running=_scheduler.is_running if _scheduler else False,
        )
    except Exception as e:
        logger.error("Health check failed", error=str(e))
        raise HTTPException(status_code=503, detail=str(e))


@app.get("/sync-status", response_model=StatusResponse, tags=["Health"])
async def get_sync_status():
    """
    Get current synchronization status for all models.

    Returns sync state for each configured model including:
    - Last sync date and ID
    - Record count
    - Status (pending, running, completed, failed)
    """
    try:
        engine = get_sync_engine()
        status = engine.get_sync_status()

        return StatusResponse(
            status="ok",
            models=status.get("models", []),
            total_models=status.get("total_models", 0),
            synced_models=status.get("synced_models", 0),
            scheduler_running=_scheduler.is_running if _scheduler else False,
        )
    except Exception as e:
        logger.error("Status check failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sync-history", response_model=SyncHistoryResponse, tags=["Health"])
async def get_sync_history(
    model_name: Optional[str] = Query(None, description="Filter by model name"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum records to return"),
):
    """
    Get sync history records.

    Returns history of sync operations including:
    - Start and end times
    - Duration
    - Records processed (inserted, updated, deleted)
    - Error count and messages
    - Odoo/PostgreSQL counts before and after
    """
    try:
        pg = PostgresClient()
        records = pg.get_sync_history(model_name=model_name, limit=limit)
        pg.close()

        return SyncHistoryResponse(
            records=records,
            total=len(records),
        )
    except Exception as e:
        logger.error("History check failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/sync-audit", response_model=SyncAuditResponse, tags=["Health"])
async def get_sync_audit(
    model_name: Optional[str] = Query(None, description="Filter by model name"),
    limit: int = Query(100, ge=1, le=1000, description="Maximum records to return"),
):
    """
    Get sync audit records comparing Odoo and PostgreSQL counts.

    Returns audit records showing:
    - Odoo record count
    - PostgreSQL record count
    - Difference between counts
    - Whether counts match (is_synced)
    - Audit timestamp
    """
    try:
        # This would query the sync_audit table
        # For now, return placeholder structure
        return SyncAuditResponse(
            records=[],
            total=0,
        )
    except Exception as e:
        logger.error("Audit check failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# Sync Endpoints
# ============================================

@app.post("/sync", response_model=SyncResponse, tags=["Sync"])
async def trigger_sync(
    request: SyncRequest,
    background_tasks: BackgroundTasks,
):
    """
    Trigger a synchronization operation.

    Can run synchronously (blocking) or be queued as a background task.
    Use full_sync=true for complete data synchronization.
    """
    try:
        logger.info("Sync triggered via API", full_sync=request.full_sync)

        engine = get_sync_engine()

        # Run sync
        results = engine.sync_all(
            full_sync=request.full_sync,
            model_names=request.models,
        )

        # Format results
        result_list = []
        for result in results:
            result_list.append({
                "model": result.model_name,
                "table": result.table_name,
                "success": result.success,
                "records_synced": result.records_synced,
                "records_inserted": result.records_inserted,
                "records_updated": result.records_updated,
                "records_deleted": result.records_deleted,
                "duration_seconds": result.duration_seconds,
                "errors": result.errors,
            })

        successful = sum(1 for r in results if r.success)
        status = "completed" if successful == len(results) else "partial"

        return SyncResponse(
            status=status,
            message=f"Synced {successful}/{len(results)} models successfully",
            results=result_list,
        )

    except Exception as e:
        logger.error("Sync failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sync/{model_name}", response_model=SyncResponse, tags=["Sync"])
async def sync_model(model_name: str, full_sync: bool = False):
    """
    Synchronize a specific model.

    Args:
        model_name: Odoo model technical name (e.g., 'res.partner').
        full_sync: Whether to perform full sync.
    """
    try:
        engine = get_sync_engine()

        # Find model config
        model_config = None
        for config in engine.config.models:
            if config.odoo_model == model_name:
                model_config = config
                break

        if not model_config:
            raise HTTPException(
                status_code=404,
                detail=f"Model '{model_name}' not found in configuration",
            )

        result = engine.sync_model(model_config, full_sync=full_sync)

        return SyncResponse(
            status="completed" if result.success else "failed",
            message=f"Synced {result.records_synced} records",
            results=[{
                "model": result.model_name,
                "table": result.table_name,
                "success": result.success,
                "records_synced": result.records_synced,
                "records_inserted": result.records_inserted,
                "records_updated": result.records_updated,
                "records_deleted": result.records_deleted,
                "duration_seconds": result.duration_seconds,
                "errors": result.errors,
            }],
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error("Model sync failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/reset", response_model=SyncResponse, tags=["Sync"])
async def reset_sync_state(request: ResetRequest):
    """
    Reset synchronization state for specified models.

    After reset, incremental sync will re-sync all records.
    """
    try:
        pg = PostgresClient()
        state_mgr = StateManager(pg)
        state_mgr.initialize()

        for model_name in request.models:
            state_mgr.reset_model_state(model_name)

        pg.close()

        return SyncResponse(
            status="completed",
            message=f"Reset sync state for {len(request.models)} models",
        )

    except Exception as e:
        logger.error("Reset failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/models", tags=["Configuration"])
async def list_models():
    """
    List all configured models.

    Returns model configurations including Odoo model name,
    PostgreSQL table name, field mappings, and deletion strategy.
    """
    try:
        engine = get_sync_engine()

        models = []
        for config in engine.config.models:
            models.append({
                "odoo_model": config.odoo_model,
                "postgres_table": config.postgres_table,
                "description": config.description,
                "deletion_strategy": config.deletion_strategy,
                "batch_size": config.batch_size or engine.config.default_batch_size,
                "field_count": len(config.fields),
                "fields": [
                    {
                        "odoo_field": f.odoo_field,
                        "postgres_column": f.postgres_column,
                        "postgres_type": f.postgres_type,
                        "primary_key": f.primary_key,
                        "is_sync_date": f.is_sync_date,
                        "field_type": f.field_type,
                        "is_foreign_key": f.is_foreign_key,
                        "indexed": f.indexed,
                    }
                    for f in config.fields
                ],
            })

        return {"models": models, "total": len(models)}

    except Exception as e:
        logger.error("List models failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


# ============================================
# Scheduler Endpoints
# ============================================

@app.post("/scheduler/start", tags=["Scheduler"])
async def start_scheduler(interval_minutes: int = 15):
    """
    Start the synchronization scheduler.

    Args:
        interval_minutes: Incremental sync interval in minutes.
    """
    global _scheduler

    if _scheduler and _scheduler.is_running:
        return {"status": "already_running", "message": "Scheduler is already running"}

    _scheduler = SyncScheduler(incremental_interval_minutes=interval_minutes)
    _scheduler.start(run_immediately=True)

    return {"status": "started", "message": f"Scheduler started with {interval_minutes}min interval"}


@app.post("/scheduler/stop", tags=["Scheduler"])
async def stop_scheduler():
    """Stop the synchronization scheduler."""
    global _scheduler

    if not _scheduler or not _scheduler.is_running:
        return {"status": "not_running", "message": "Scheduler is not running"}

    _scheduler.stop()
    _scheduler = None

    return {"status": "stopped", "message": "Scheduler stopped"}


@app.get("/scheduler/status", tags=["Scheduler"])
async def scheduler_status():
    """Get scheduler status and next run times."""
    if not _scheduler:
        return {
            "running": False,
            "next_runs": [],
        }

    next_runs = _scheduler.get_next_run_times(5)
    return {
        "running": _scheduler.is_running,
        "next_runs": [
            {"datetime": run.isoformat(), "timestamp": run.timestamp()}
            for run in next_runs
        ],
    }


@app.get("/validate", tags=["Configuration"])
async def validate_configuration():
    """
    Validate the current configuration.

    Checks that all configured models and fields exist in Odoo.
    """
    try:
        engine = get_sync_engine()
        errors = engine.validate_configuration()

        return {
            "valid": len(errors) == 0,
            "errors": errors,
        }

    except Exception as e:
        logger.error("Validation failed", error=str(e))
        raise HTTPException(status_code=500, detail=str(e))


# Import after auth helpers are defined so the Control Tower router can reuse
# the existing signed dashboard session without a second authentication layer.
from src.control_tower.router import router as control_tower_router  # noqa: E402

app.include_router(control_tower_router)


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(
        "src.api:app",
        host="0.0.0.0",
        port=8000,
        reload=True,
    )



